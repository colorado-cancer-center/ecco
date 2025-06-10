#!/usr/bin/env python

from collections import Counter, defaultdict
import csv
import json
import os
from pprint import pprint
from urllib.parse import quote

import openpyxl

import click
import requests as rq
from tqdm import tqdm

# from slugify import slugify

from diskcache import Cache

# ==============================================================================
# === Metadata, extracted from hardcoded locations.json
# ==============================================================================

# this dict maps keys from the source data to their keys in the final
# locations-data.json file.the final keys are referenced in the hardcoded
# file locations.json.
SRC_LABELS_TO_LOCATIONS_DATA = {
    "ccrm": {
        'AccessoriesAids_29Jul21': 'accessories-aids-29-jul-21',
        'AllSupportServices_15Jun2021': 'all-support-services-15-jun-2021',
        'AllTreatment_15Jun2021': 'all-treatment-15-jun-2021',
        'CancerEd_29Jul221': 'cancer-ed-29-jul-221',
        'CCSP_15Jun2021': 'ccsp-15-jun-2021',
        'CHC_15Jun2021': 'chc-15-jun-2021',
        'FinancialCoun_15Jun2021': 'financial-coun-15-jun-2021',
        'Fitness_29Jul21': 'fitness-29-jul-21',
        'GeneticCoun_15Jun2021': 'genetic-coun-15-jun-2021',
        'HomeFamilyCare_29Jul21': 'home-family-care-29-jul-21',
        'Hospice_15Jun2021': 'hospice-15-jun-2021',
        'IHS_15Jun2021': 'ihs-15-jun-2021',
        'Lodging_15Jun2021': 'lodging-15-jun-2021',
        'MedOnc_15Jun2021': 'med-onc-15-jun-2021',
        'Nutrition_15Jun2021': 'nutrition-15-jun-2021',
        'Palliative_15Jun2021': 'palliative-15-jun-2021',
        'PedOnc_15Jun2021': 'ped-onc-15-jun-2021',
        'ProgramsEvents_29Jul21': 'programs-events-29-jul-21',
        'PTRehab_15Jun2021': 'pt-rehab-15-jun-2021',
        'RadOnc_15Jun2021': 'rad-onc-15-jun-2021',
        'RHC_15Jun2021': 'rhc-15-jun-2021',
        # 'SupportCoun_15Jun2021': 'support-coun-15-jun-2021',
        'SupportGrp_15Jun2021': 'support-grp-15-jun-2021',
        'SurgOnc_15Jun2021': 'surg-onc-15-jun-2021',
        'Survivorship_15Jun2021': 'survivorship-15-jun-2021',
        'TitleX_15Jun2021': 'title-x-15-jun-2021',
        'Transportation_29Jul21': 'transportation-29-jul-21',
        'Trials_15Jun2021': 'trials-15-jun-2021',
        'WWC_15Jun2021': 'wwc-15-jun-2021',
    },
    "cif": {
        'Colon & Rectal Surgeon': 'colon-rectal-surgeon',
        'FQHC': 'fqhc',
        'Gastroenterology': 'gastroenterology',
        'HPSA Correctional Facility': 'hpsa-correctional-facility',
        'HPSA Federally Qualified Health Center Look A Like': 'hpsa-federally-qualified-health-center-look-a-like',
        'HPSA Indian Health Service, Tribal Health, and Urban Indian Health Organizations': 'hpsa-indian-health-service-tribal-health-and-urban-indian-health-organizations',
        'HPSA Other Facility': 'hpsa-other-facility',
        'HPSA Rural Health Clinic': 'hpsa-rural-health-clinic',
        'Lung Cancer Screening': 'lung-cancer-screening',
        'Mammography': 'mammography',
        'Obstetrics & Gynecology': 'obstetrics-gynecology',
        'Radiation Oncology': 'radiology',
        'Superfund Site': 'superfund-site',
        'Toxic Release Inventory Facility': 'toxic-release-inventory-facility',
    },
    "legislative": {
        'Senate Districts': 'senate-districts',
        'House Districts': 'house-districts',
        'Congressional Districts': 'congressional-districts',
    }
}


# ==============================================================================
# === Geocoding
# ==============================================================================

MAPS_API_KEY = os.environ['MAPS_API_KEY']
MAPS_CCTLD_REGION = 'en'

MAPS_GEOCODE_URL_PREFIX = 'https://maps.googleapis.com/maps/api/geocode/json'

cache = Cache(".geocoding")

@cache.memoize()
def geocode(address):
    """
    Geocodes an address using Google Maps API, returns
    a tuple of (longitude, latitude).
    """
    url = f'{MAPS_GEOCODE_URL_PREFIX}?region={MAPS_CCTLD_REGION}&address={quote(address)}&key={MAPS_API_KEY}'
    response = rq.get(url)
    data = response.json()

    try:
        point = data["results"][0]["geometry"]["location"]

    except Exception as ex:
        print(f"URL: {url}")
        pprint(data)
        raise ex

    return (float(point["lng"]), float(point["lat"]))


# ==============================================================================
# === Source-specific Extractions
# ==============================================================================

def entries_in_locations(locations):
    possible_loc_names = []

    for _category, items in locations.items():
        for item in items.values():
            possible_loc_names.append(item)

    return possible_loc_names

def extract_cif_features(cif_locations_csv):
    # collect all features into a dict of the form {location_type: [features]}
    features = defaultdict(list)

    # get the number of rows in cif_locations_csv by counting the lines, then
    # reset the file pointer so we can read the actual data later
    cif_locations_csv.seek(0)
    num_cif_rows = sum(1 for _ in cif_locations_csv) - 1
    cif_locations_csv.seek(0)

    # read in the CIF locations CSV
    # header:
    # Type,Name,Address,State,Phone_number,Notes,latitude,longitude,FIPS
    for row in tqdm(csv.DictReader(cif_locations_csv), total=num_cif_rows):
        loc_type = row["Type"].strip()
        if (
            row["State"].strip() != "CO" or 
            loc_type not in SRC_LABELS_TO_LOCATIONS_DATA['cif']
        ):
            continue

        # figure out the key under which we'll associate it with the locations metadata
        dest_type = SRC_LABELS_TO_LOCATIONS_DATA['cif'][loc_type]

        # check if the latitude and longitude are already present
        # and parseable as floats; if not, attempt to derive them by geocoding
        # the address
        try:
            lat_float = float(row["latitude"])
            lon_float = float(row["longitude"])
        except (ValueError, KeyError):
            lat_float = None
            lon_float = None

        try:
            if lat_float is None or lon_float is None:
                # print(f"Geocoding {row['Name']}, address: {row['Address']}...")
                point = geocode(row["Address"])
                lat_float, lon_float = point
            
            features[dest_type].append({
                "type": "Feature",
                "properties": {
                    "name": row["Name"],
                    "address": row["Address"],
                    "phone": row["Phone_number"],
                    "notes": row["Notes"],
                },
                "geometry": {
                    "type": "Point",
                    "coordinates": [lon_float, lat_float],
                }
            })

        except KeyError as ex:
            print(f"Error processing {row['Type']} -> {row['Name']}: missing key {ex}, continuing...")
            raise ex

    # return the features with each category wrapped in a FeatureCollection
    return {
        k: {"type": "FeatureCollection", "features": v}
        for k, v in features.items()
    }

def extract_ccrm_features(ccrm_json):
    # collect all features into a dict of the form {location_type: [features]}
    features = defaultdict(list)

    ccrm = json.load(ccrm_json)
    operational_layers = ccrm['operationalLayers']

    total_errors = 0
    error_field_types = Counter()

    # retain only the layers that have a title that matches a location name
    for layer in tqdm(operational_layers):
        layer_title = layer['title'].strip()
        print(f"Checking {layer_title}...")
        if layer_title not in SRC_LABELS_TO_LOCATIONS_DATA['ccrm']:
            continue
        
        dest_type = SRC_LABELS_TO_LOCATIONS_DATA['ccrm'][layer_title]

        # the features we actually want are in the following path under
        # the object with a matching 'title' value:
        # featureCollection.layers[0].featureSet.features[].attributes
        for feature in layer['featureCollection']['layers'][0]['featureSet']['features']:
            attrs = feature['attributes']

            # check if it has ADDRESS1 and ADDRESS2 rather than ADDRESS
            if 'ADDRESS' not in attrs and 'ADDRESS1' in attrs and 'ADDRESS2' in attrs:
                attrs['ADDRESS'] = " ".join(
                    x for x in [attrs['ADDRESS1'], attrs['ADDRESS2']]
                    if x and x.strip() != ""
                )

            # check if it has ZIPCODE instead of ZIP
            if 'ZIP' not in attrs and 'ZIPCODE' in attrs:
                attrs['ZIP'] = str(attrs['ZIPCODE'])

            try:
                candidate = {
                    "type": "Feature",
                    "properties": {
                        "name": attrs['NAME'],
                        "org": attrs.get('ORGANIZATI'),
                        "link": attrs['WEBSITE'],
                        "address": " ".join([
                            attrs['ADDRESS'], attrs['CITY'], attrs['STATE'], attrs['ZIP']
                        ]),
                        "phone": attrs['PHONE']
                    },
                    "geometry": {
                        "type": "Point",
                        "coordinates": [float(attrs['LONGITUDE']), float(attrs['LATITUDE'])],
                    }
                }

                # retain only properties that are not empty strings
                candidate['properties'] = {
                    k: v for k, v in candidate['properties'].items() if v and v.strip()
                }

                features[dest_type].append(candidate)

            except KeyError as ex:
                total_errors += 1
                print(f"Error processing {layer_title} -> {attrs['NAME']}: {ex}, continuing...")
                error_field_types[str(ex)] += 1
                continue

    if total_errors > 0:
        print(f"Total errors: {total_errors}")
        print("Error field types:")
        pprint(error_field_types)
        print()

    # return the features with each category wrapped in a FeatureCollection
    return {
        k: {"type": "FeatureCollection", "features": v}
        for k, v in features.items()
    }


def extract_legislative_features(
        house_geojson, senate_geojson, congressional_geojson,
        house_senate_reps_excel, congressional_reps_csv
):
    # collect all features into a dict of the form {location_type: [features]}
    features = defaultdict(list)

    # read in geojson files for each legislative district type
    geojson_map = {
        "house-districts": json.load(house_geojson),
        "senate-districts": json.load(senate_geojson),
        "congressional-districts": json.load(congressional_geojson),
    }

    # read in the house and senate XLSX file and create
    # a dict that maps district numbers to the representative's info
    all_reps = {
        "house-districts": {},
        "senate-districts": {},
        "congressional-districts": {},
    }

    # use openpyxl to load in the excel sheet
    try:
        # DATA NOTES:
        # - there's a field named 'Committee Membership' in the sheet, but
        #   not currently surfaced in the UI.
        # - there's an extra column in "Senate Members" with no header that
        #   has notes in it, e.g. "resigned on <date>"
        wb = openpyxl.load_workbook(house_senate_reps_excel)

        # map chambers to worksheets
        chambers_to_worksheets = {
            "senate-districts": wb["Senate Members"],
            "house-districts": wb["House Members"]
        }

        # iterate over each known worksheet in the model
        for chamber, ws in chambers_to_worksheets.items():
            # read the first row as the header
            headers = [cell.value for cell in ws[1]]

            # create dict per row with keys from the header
            reps = [
                dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True)
            ]

            # populate the house_senate_reps dict, keyed by district
            for rep in reps:
                district = int(rep["District"])
                all_reps[chamber][district] = {
                    "district": district,
                    "representative": f"{rep['First Name']} {rep['Last Name']}",
                    "party": rep["Party Affiliation"],
                    "phone": rep["Work Phone"],
                    "email": rep["Email"],
                }
    finally:
        wb.close()

    # read in the congressional reps CSV file and create a dict
    # that maps district numbers to the representative's info
    for row in csv.DictReader(congressional_reps_csv):
        district = int(row["DistrictNum"])
        last_name, first_name = row["Member"].split(", ")

        all_reps["congressional-districts"][district] = {
            "district": district,
            "representative": f"{first_name} {last_name}",
            "party": row["Affiliation"],
            "website": row["Website"]
        }

    # at this point, all_reps is populated with all
    # the house, senate, and congressional representatives
    pprint(all_reps)

    # iterate over each district type and its features
    for district_type, geojson in geojson_map.items():
        for feature in geojson["features"]:
            try:
                # get the district number
                district_number = feature["properties"]["District"]

                # get the representative info for this district
                rep_info = all_reps[district_type][district_number]

                # add the representative info to the feature
                feature["properties"].update(rep_info)

                # add the feature to the list of features
                features[district_type].append(feature)
            except KeyError as ex:
                print(f"Error in {district_type} -> {district_number}: missing key {ex}")
                raise ex

    # return the features with each category wrapped in a FeatureCollection
    return {
        k: {"type": "FeatureCollection", "features": v}
        for k, v in features.items()
    }


# ==============================================================================
# === Entrypoint
# ==============================================================================

def _flatten_dict(in_dict):
    """
    Recursively flatten in_dict to produce
    just a list of values
    """
    out_list = []

    for v in in_dict.values():
        if isinstance(v, dict):
            out_list.extend(_flatten_dict(v))
        else:
            out_list.append(v)

    return out_list

@click.command()
@click.argument('locations_json', type=click.File('r'))
@click.argument('ccrm_json', type=click.File('r'))
@click.argument('cif_locations_csv', type=click.File('r'))
@click.argument('house_geojson', type=click.File('r'))
@click.argument('senate_geojson', type=click.File('r'))
@click.argument('congressional_geojson', type=click.File('r'))
@click.argument('house_senate_reps_excel', type=click.File('rb'))
@click.argument('congressional_reps_csv', type=click.File('r'))
@click.option('--output', help='Output file', type=click.File('w'), default='-')
def produce_features(
    locations_json, ccrm_json, cif_locations_csv,
    house_geojson, senate_geojson, congressional_geojson,
    house_senate_reps_excel, congressional_reps_csv,
    output
):
    # create dict that will store features for each location type
    features = defaultdict(list)

    # ------------------------------------
    # --- 1. process sources
    # --- 1a: extract possible location names from locations.json
    # --- 1b: extract features from the CIF locations CSV
    # --- 1c: extract features from the CCRM JSON
    # --- 1d: extract features from the legislative data
    # ------------------------------------
    
    # 1a. open locations.json and read all the values into a list
    defined_locs = set(entries_in_locations(json.load(locations_json)))
    expected_locs = set(_flatten_dict(SRC_LABELS_TO_LOCATIONS_DATA))

    # print items that are in defined_locs but not in expected_locs
    print("Defined but not expected:")
    pprint(defined_locs - expected_locs)
    # print items that are in expected_locs but not in defined_locs
    print("Expected but not defined:")
    pprint(expected_locs - defined_locs)

    assert set(defined_locs) == set(_flatten_dict(SRC_LABELS_TO_LOCATIONS_DATA))

    # 1b. extract features from the CIF locations CSV
    cif_features = extract_cif_features(cif_locations_csv)
    features.update(cif_features)
    
    # 1c. read the operationalLayers key from the CCRM JSON
    # and extract features from it
    ccrm_features = extract_ccrm_features(ccrm_json)
    features.update(ccrm_features)

    # 1d. combine legislative geojson files, annotate w/reps for each
    # location
    legislative_features = extract_legislative_features(
        house_geojson, senate_geojson, congressional_geojson,
        house_senate_reps_excel, congressional_reps_csv
    )
    features.update(legislative_features)

    # write the relevant layers to the output file
    json.dump(features, output, indent=2)

if __name__ == '__main__':
    produce_features()
