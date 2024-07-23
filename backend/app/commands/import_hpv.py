#!/usr/bin/env python

from pprint import pprint
import sys

sys.path.append("/app")

import click
import asyncio

import openpyxl
from openpyxl.utils import get_column_letter

from tqdm import tqdm
from sqlmodel import delete, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import AsyncSession

from db import engine

from models.geom import County
from models.hpv import (
    HPVCounty
)


class ModelNotProvidedException(Exception):
    """
    Raised when an input file is found that has no associated
    model class.
    """
    pass

class UnexpectedSheetException(Exception):
    """
    Raised when the sheet in the input file is not what was expected
    (i.e., it didn't match the name in HPV_SHEETS_META for that index)
    """
    pass

async def import_measure(model, measure, rows, session):
    """
    Imports a table of measurements into 'model' in the database.

    Note that the "rows" object is created in the import_radon function; it's
    not just the original row from the spreadsheet. We expect that each row
    object has the following keys: "fips", "county", and "value".

    We reuse the value column as the name of the measure, so the measures will
    come from HPV_MEASURES below.
    """

    # read each row, creating an object from it and adding it to the list
    obj_list = []
    for row in rows:
        # create a new object
        obj_list.append(model(**{
            "FIPS" : row["fips"],
            "County" : row["county"],
            "State" : "Colorado",
            "measure" : measure,
            "value": row["value"],
            "sex": row["sex"]
        }))

    # bulk insert all objects
    session.add_all(obj_list)

# columns in both county and tract sheets that we'll import as measures
HPV_MEASURES = [
    "NTests", # Total Tests
    "NTestsover4", # Total Test over 4 pCi/L
    "PctOver4", # Percentage of Tests over 4 pCi/L
]

# the spreadsheet lists gender under the "Vaccine" column, so
# we need to map those values to the genders "All", "Male", "Female"
VACCINE_TO_GENDER = {
    "Patients (Any Gender) UTD for HPV": "All",
    "Patients (Females Only) UTD for HPV": "Female",
    "Patients (Males Only) UTD for HPV": "Male"
}

async def import_hpv_data(sheets, delete_before_import=True):
    """
    Given a set of sheets with CDPHE HPV vaccination info, imports them one by one into the database.

    Expects the sheets to have the following columns:
    - County
    - Vaccine
    - Colorado NIS Rate
    - US NIS Rates
    - HP2020 Goals
    - Period
    - Percent Group (redundant with up-to-date percent; just groups them into buckets)
    - Up-To-Date Percent (used as the measure's value)
    - State Rate

    This data can be found at the following website:
    https://cohealthviz.dphe.state.co.us/t/DCEED_Public/views/CountyRateMaps-Storyboard/CountyRateMapsCombined?%3Aembed=y&%3AisGuestRedirectFromVizportal=y

    Iterates over each sheet, zipping it with HPV_SHEETS_META to get
    metadata about the sheet. Then, for each column in HPV_SHEETS_META,
    it imports that column's value, using the column name as the measure name.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    # step 1. convert input excel using the column mapper

    # for each file, import it into the database
    async with async_session() as session:
        # delete all existing entries in the target models
        if delete_before_import:
            for model in (HPVCounty, ):
                result = await session.execute(delete(model))
                tqdm.write(f" - Deleted {result.rowcount} from {model.__name__}")

        # get a dict of county names to FIPS mappings so we don't have to
        # do a ton of queries
        county_fips_results = await session.execute(select(County.county, County.us_fips))
        county_to_fips = {
            row[0]: row[1]
            for row in county_fips_results.all()
        }

        # read in the excel file via openpyxl
        for sheet in sheets:
            tqdm.write(f"* About to process {sheet}...")

            try:
                wb = openpyxl.load_workbook(sheet)
                ws = wb.active

                # create a dictionary of column names
                col_names = [
                    x[0].strip()
                    for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
                    if x[0] is not None
                ]
                col_indices = { col: idx for idx, col in enumerate(col_names) }

                # import measure for the current sheet
                value_col = "Up-To-Date Percent"
                tqdm.write(f"* Processing {ws.title}, {value_col}...")

                # read the column, normalizing sheet cols to internal names
                rows = [
                    {
                        "fips": county_to_fips[row[col_indices["County"]].upper()],
                        "county": row[col_indices["County"]],
                        "value": row[col_indices[value_col]] / 100.0,
                        "sex": VACCINE_TO_GENDER.get(row[col_indices["Vaccine"]])
                    }
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[col_indices[value_col]] is not None
                ]
                
                await import_measure(
                    model=HPVCounty,
                    measure=value_col,
                    rows=rows,
                    session=session
                )

                # commit session at the end
                await session.commit()

            finally:
                wb.close()

@click.command()
@click.argument('excel-sheets', type=str, nargs=-1)
def main(excel_sheets):
    asyncio.run(import_hpv_data(excel_sheets))

if __name__ == '__main__':
    main()
