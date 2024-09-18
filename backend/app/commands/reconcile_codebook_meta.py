#!/usr/bin/env python

from collections import defaultdict
import csv
import json
from pathlib import Path
from pprint import pprint
import re
import sys
from typing import Optional

sys.path.append("/app")

import click
import asyncio

import openpyxl
from openpyxl.utils import get_column_letter

from tqdm import tqdm

from models.cif_meta import CIF_MEASURE_DESCRIPTIONS
from models.base import MeasureUnit

# maps the 'Table' column to our internal measure categories
TABLE_TO_META_CATEGORIES = {
    "Sociodemographic": "sociodemographics",
    "Economic & Insurance": "economy",
    "Housing & Transportation": "housingtrans",
    "Disparities": "disparities",
    "Risk Factor & Screening": "rfandscreening",
    "Food Desert": "fooddesert",
    "Environment": "environment",
    # "Cancer Incidence": "cancermortality", # we use SCP cancer data, so we can ignore these
    # "Cancer Mortality": "cancermortality", # we use SCP cancer data, so we can ignore these
    # "Facilities & Providers": "", # we have no equivalent, since we deal with locations separately
}

# some variable names differ between the input data and the codebook
# for example, the source data includes a measure named "Other_Races",
# whereas the codebook refers to it as "Other Races", with a space.
# this mapping allows us to reconcile these differences
VARIABLE_NAME_TO_MEASURE = {
    "Other Races": "Other_Races",
    "Urban Percentage": "Urban_Percentage",
    "Lack English Prof": "Lack_English_Prof",
    "Cancer Prevalence": "Cancer_Prevalence",
}

# the codebook is not up-to-date with the actual input data; for example,
# "Gini Coefficient" is now under "disparities", whereas the codebook lists
# it under "Economic & Insurance". This mapping allows us to remap these
# categories to their correct values.
# the format is (old_category, measure_name): (new_category, measure_name)
CATEGORY_REMAPS = {
    ("Economic & Insurance", "Gini Coefficient"): ("Disparities", "Gini Coefficient"),
    ("Economic & Insurance", "Vacancy Rate"): ("Housing & Transportation", "Vacancy Rate"),
    ("Economic & Insurance", "No Vehicle"): ("Housing & Transportation", "No Vehicle"),
    ("Economic & Insurance", "Rent Burden (40% Income)"): ("Housing & Transportation", "Rent Burden (40% Income)"),
}

# yet another place where the codebook is incorrect: the 2024-09
# release has apparently removed Met_Cervical_Screening from
# screening & risk factors, so we need to skip it. the list below
# contains a set of skipped measures
SKIPPED_MEASURES = {
    "Met_Cervical_Screen",
    "Kidney_Disease",
    "PWS_Violations_Since_2016",
    "Air_Toxics_Cancer",
    "Air_Toxics_Resp",
}

class TypeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, set):
            return list(obj)
        if isinstance(obj, MeasureUnit):
            return re.match(r"^<([^:]+): .*>$", repr(obj)).group(1)
        return json.JSONEncoder.default(self, obj)

def stats_tables(stats_dir, use_cache=True):
    """
    Given a directory of stats files, reads in all of the
    files and returns a dict of measure category names
    and the measures they contain.
    """

    print("* Loading stats files...")

    # first, if we're using the cache, check that it's there
    cache_file = Path(stats_dir) / "measure_cache.json"
    if use_cache:
        if cache_file.exists():
            try:
                print("...found cache file and use_cache is true, loading from cache")
                with open(cache_file, 'r') as f:
                    return json.load(f)
            except json.JSONDecodeError:
                print("...cache file is invalid, regenerating")
    
    # glob all *_.csv files in the directory
    # read in each file and extract the column names
    # return a dict of category names to measures

    files = list(Path(stats_dir).glob("*_long_*.csv"))

    measure_categories = defaultdict(set)

    for file in tqdm(files):
        # use csv.DictReader to read in each file
        # for cancer models (i.e., files in which the name 'cancer' appears), we
        # extract the column named 'Site' as the measure. for the rest, it's
        # 'measure'

        # extract the category from a filename like us_environment_tract_08-30-2024.csv
        category = re.match(r"^us_(\w+)_(county|tract)", file.name).group(1)

        with open(file, 'r') as f:
            reader = csv.DictReader(f)

            for row in reader:
                measure_categories[category].add(
                    row['Site'] if 'cancer' in file.name else
                    row['measure']
                )

    # write out the cache for next time
    with open(cache_file, 'w') as f:
        return json.dump(measure_categories, fp=f, cls=TypeEncoder, indent=4)
    
    return measure_categories


async def compare_codebook(excel_path, stats_dir=None, update_labels=False):
    """
    Given a path to the CiF codebook, loads it and
    compares it to the CiF metadata. Produces an annotated
    version of the metadata dict, filled in from the codebook.

    If a stats directory is provided, also uses a CiF data
    distribution to identify which columns are actually used
    in that month's data.
    """

    measure_categories = (
        stats_tables(stats_dir) if stats_dir is not None else None
    )

    # read in the excel file via openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path)
        # get the first sheet
        ws = wb.worksheets[0]

        # create a dictionary of column names
        col_names = [
            x[0].strip()
            for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
            if x[0] is not None
        ]
        col_indices = { col: idx for idx, col in enumerate(col_names) }

        # add whitespace
        print()

        # run through each measure and compare it to the meta
        for row in ws.iter_rows(2, ws.max_row, 1, ws.max_column, values_only=True):
            # --------------------------------------------
            # - 1. extract values from columns
            # --------------------------------------------

            # get the measure name
            table_name = row[col_indices["TABLE"]]
            measure_name = row[col_indices["VARIABLE NAME"]]
            measure_label = row[col_indices["VARIABLE DEFINITION"]]
            source = row[col_indices["SOURCE"]]
            source_url = row[col_indices["URL"]]

            # --------------------------------------------
            # - 2. skip, remap categories and measures
            # --------------------------------------------

            # skip categories we're not reconciling, e.g. cancer-related ones
            if table_name not in TABLE_TO_META_CATEGORIES:
                print(f"~ Skipping category {table_name} w/measure {measure_name}")
                continue

            # remap categories if necessary
            if (table_name, measure_name) in CATEGORY_REMAPS:
                table_name, measure_name = CATEGORY_REMAPS[(table_name, measure_name)]

            # map their table names into our internal categories
            category_name = TABLE_TO_META_CATEGORIES[table_name]

            # map measures to our internal names, if necessary
            if measure_name in VARIABLE_NAME_TO_MEASURE:
                measure_name = VARIABLE_NAME_TO_MEASURE[measure_name]

            # if the measure isn't found in the meta, first replace spaces with underscores
            if measure_name not in CIF_MEASURE_DESCRIPTIONS[category_name]:
                measure_name = measure_name.replace(" ", "_")

            # skip measures that are no longer present in the data
            if measure_name in SKIPPED_MEASURES:
                print(f"- Explicitly skipped measure {measure_name} as it is no longer present in the data")
                continue

            # --------------------------------------------
            # - 3. apply patches
            # --------------------------------------------

            # retrieve the current metadata for this category and measure
            try:
                cur_object = CIF_MEASURE_DESCRIPTIONS[category_name][measure_name]
            except KeyError as ex:
                print(f"[!] Unable to find measure {measure_name} in category {category_name}")

                # see if we can find it elsewhere in the meta
                for category in measure_categories:
                    if measure_name in measure_categories[category]:
                        print(f"...but found measure {measure_name} in category {category}")
                raise ex

            # check for differences between the codebook and the metadata
            if cur_object["label"] != measure_label:
                if update_labels:
                    print(f"x Changing label for '{measure_name}' from '{cur_object['label']}' to '{measure_label}'")
                    cur_object["label"] = measure_label
                else:
                    print(f"? Found difference in '{measure_name}'; + meta: '{cur_object['label']}', - codebook: '{measure_label}'")

            # merge in changes, preferring ours to theirs
            CIF_MEASURE_DESCRIPTIONS[category_name][measure_name] = {
                **{
                    "label": measure_label,
                    "source": source,
                    "source_url": source_url,
                },
                **cur_object
            }

        # pprint(CIF_MEASURE_DESCRIPTIONS)
        print()
        print(
            """-----------------------------------------\n"""
            """- Final metadata\n"""
            """-----------------------------------------\n"""
        )
        print(
            json.dumps(CIF_MEASURE_DESCRIPTIONS, indent=4, cls=TypeEncoder)
        )

    finally:
        wb.close()

@click.command()
@click.argument('excel-path', type=str)
@click.argument('stats-dir', type=str, required=False)
def main(excel_path, stats_dir=None):
    """
    Reads in the CancerInFocus codebook and compares it to the
    CiF metadata dict from backend.models.cif_meta. Produces a version
    of the metadata dict in which the codebook values are used
    when present, and if not the CiF metadata values are used.

    Optionally uses a data distribution from CiF to attempt to identify
    which columns are actually used, too.
    """
    asyncio.run(compare_codebook(excel_path, stats_dir))

if __name__ == '__main__':
    main()
