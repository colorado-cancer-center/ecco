#!/usr/bin/env python

import sys

sys.path.append("/app")

import click
import asyncio

import openpyxl

from tqdm import tqdm
from sqlmodel import delete
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import AsyncSession

from db import engine

from models.ccc_state_stats import (
    StateCancerIncidenceStats,
    StateCancerMortalityStats,
    StateSociodemographicStats
)


# maps worksheets in the input file to the model they should be imported into
CANCER_COL_MODEL_MAP = {
    "Cancer": "site",
    "Race/Ethnicity": "race",
    "Sex": "sex",
    "Age": "age",
    "Stage": "stage",
    "StateAvg": "state_avg",
    "USAvg": "us_avg",
}

SHEETS_TO_MODELS = {
    "ECCO_IncAverages": {
        "model": StateCancerIncidenceStats,
        "colmap": CANCER_COL_MODEL_MAP
    },
    "MortAverages": {
        "model": StateCancerMortalityStats,
        "colmap": CANCER_COL_MODEL_MAP
    },
    "ScreeningRiskFactors": {
        "model": StateSociodemographicStats,
        "constants": {
            "measure_category": "Screening & Risk Factors",
        },
        "colmap": {
            "Screening & Risk Factors": "measure",
            "Data Source": "source",
            "StateAvg": "state_avg",
        }
    },
    "Radon": {
        "model": StateSociodemographicStats,
        "constants": {
            "measure_category": "Radon Exposure",
        },
        "colmap": {
            "Radon Exposure": "measure",
            "Data Source": "source",
            "StateAvg": "state_avg",
        }
    },
}


async def import_measure(rows, model, session):
    """
    Imports a table of measurements into 'model' in the database.

    If delete_before_import is True, all existing entries in the target model
    are deleted before importing.
    """

    # read each row, creating an object from it and adding it to the list
    obj_list = []
    for row in rows:
        # create a new object
        obj_list.append(model(**row))

    # bulk insert all objects
    session.add_all(obj_list)

async def import_ccc_state_stats(excel_path, delete_before_import=True):
    """
    Imports the state average data compiled by Colorado Cancer Center.

    The data includes separate worksheets for different models, e.g.
    cancer incidence, cancer mortality, and sociodemographic data.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    # step 1. convert input excel using the column mapper

    # for each file, import it into the database
    async with async_session() as session:
        # pre-step: clear out the existing data
        if delete_before_import:
            for model in set(x["model"] for x in SHEETS_TO_MODELS.values()):
                result = await session.execute(delete(model))
                tqdm.write(f" - Deleted {result.rowcount} from {model.__name__}")

        tqdm.write(f"* About to process {excel_path}...")

        # read in the excel file via openpyxl
        try:
            wb = openpyxl.load_workbook(excel_path)

            # iterate over each worksheet in the workbook, mapping it to a model
            for ws in wb.worksheets:
                tqdm.write(f"* Processing {ws.title}...")

                model_info = SHEETS_TO_MODELS.get(ws.title)
                if model_info is None:
                    tqdm.write(f"  - No model provided for {ws.title}")
                    continue

                # extract metadata needed to create instances of each model
                model = model_info["model"]
                colmap = model_info["colmap"]
                constants = model_info.get("constants", {})

                # create a dictionary of column names
                col_names = [
                    x[0].strip()
                    for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
                    if x[0] is not None
                ]
                col_indices = { col: idx for idx, col in enumerate(col_names) }

                # read the columns
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {
                        **constants,
                        **{
                            colmap[col]: row[col_indices[col]]
                            for col in colmap
                        }
                    }
                    rows.append(row_dict)

                await import_measure(
                    rows, model, session
                )

                tqdm.write(f"...insert done, committing...")
                await session.commit()
                tqdm.write(f"done!\n")
                
        finally:
            wb.close()

@click.command()
@click.argument('excel-path', type=str)
def main(excel_path):
    asyncio.run(import_ccc_state_stats(excel_path))

if __name__ == '__main__':
    main()
