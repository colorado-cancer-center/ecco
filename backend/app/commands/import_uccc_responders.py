#!/usr/bin/env python

import sys
import asyncio
import click
import openpyxl
import csv

from itertools import groupby

from sqlalchemy.orm import sessionmaker
from sqlmodel import delete, select
from tqdm import tqdm

sys.path.append("/app")

from db import engine
from models.uccc_responders import UCCCRespondersCounty
from models.geom import County
from sqlalchemy.ext.asyncio import AsyncSession


def read_sheet(sheet_fp, sheet_name=None):
    """
    Given an Excel sheet, read the data from it and yield
    a dictionary for each row, where the keys are the column
    headers and the values are the cell values.
    """

    wb = openpyxl.load_workbook(sheet_fp)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    # get the header row
    header = [cell.value for cell in ws[1]]

    # get the data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(header, row))

async def import_into_model(session, model, label, records):
    """
    Helper to insert records into model.
    """
    
    tqdm.write(f"* Processing {label}...")
    new_recs = [
        model(**rec)
        for rec in records
    ]
    session.add_all(new_recs)

    tqdm.write(f"* Inserted {len(new_recs)} records")
    tqdm.write("")


async def import_uccc_responser_data(uccc_reponders_sheet, delete_before_import=True):
    """
    Imports health-region uccc_responser data from an Excel sheet into the database.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    if delete_before_import:
        async with async_session() as session:
            print("* Deleting existing UCCC responser data...")
            for model in (UCCCRespondersCounty,):
                result = await session.execute(delete(model))
                print(f" - Deleted {result.rowcount} from {model.__name__}")
            print("")
            await session.commit()

    # ======================================================================
    # import UCCC responders
    # ======================================================================

    async with async_session() as session:
        # the excel input consists of a single sheet with these three columns:
        # - ZipCode: the zip code of the responder
        # - County: the county of the responder
        # - Urbanicity: the categorical value "Urban" or "Rural"

        # get a dict of county names to FIPS mappings so we don't have to
        # do a ton of queries
        county_fips_results = await session.execute(
            select(County.county, County.us_fips)
        )
        county_to_fips = {
            row[0]: row[1]
            for row in county_fips_results.all()
        }


        sheet_rows = read_sheet(uccc_reponders_sheet)
        # first, sort by county so groupby works
        sorted_rows = sorted(
            sheet_rows,
            key=lambda r: (
                r["County"],
                r["Urbanicity"],
            )
        )
        # then group by county
        grouped_rows = groupby(
            sorted_rows,
            key=lambda r: (
                r["County"],
                r["Urbanicity"],
            )
        )

        instances = [
            {
                "FIPS": county_to_fips[county_name],
                "County": county_name,
                "State": "Colorado",
                "measure": "responses",
                "value": len(list(data)),
                "urbanicity": urbanicity,
            }
            for (county_name, urbanicity), data in grouped_rows
        ]

        await import_into_model(
            session, UCCCRespondersCounty, "UCCC Responders by County", instances
        )

        # unfortunately, the way factors are currently set up assumes that there
        # will be an 'all' entry in the data that aggregates the other values that
        # that factor can take on. this becomes an issue for datasets that aren't
        # explicitly broken out that way.
        # since our data lacks that (i.e. each zip code is urban/rural based on the
        # county in which it's located), we need to create those 'all' entries
        # ourselves by just adding in each county with urbanicity='All'
        await import_into_model(
            session, UCCCRespondersCounty, "UCCC Responders by County",
            ({ **inst, "urbanicity": "All" } for inst in instances)
        )

        await session.commit()

@click.command()
@click.argument('uccc-reponders-sheet', type=click.File('rb'))
def main(uccc_reponders_sheet):
    asyncio.run(import_uccc_responser_data(
        uccc_reponders_sheet
    ))

if __name__ == '__main__':
    main()
