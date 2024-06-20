#!/usr/bin/env python

import sys

sys.path.append("/app")

import click
import asyncio

import openpyxl
from openpyxl.utils import get_column_letter

from tqdm import tqdm
from sqlmodel import delete
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import AsyncSession

from db import engine

from models.radon import (
    RadonCounty, RadonTract
)


class ModelNotProvidedException(Exception):
    """
    Raised when an input file is found that has no associated
    model class.
    """
    pass

class UnexpectedSheetException(Exception):
    """
    Raised when the sheet in the input file is not what was expected
    (i.e., it didn't match the name in RADON_SHEETS_META for that index)
    """
    pass

async def import_measure(model, measure, rows, session):
    """
    Imports a table of measurements into 'model' in the database.

    Note that the "rows" object is created in the import_radon function; it's
    not just the original row from the spreadsheet. We expect that each row
    object has the following keys: "fips", "county", and "value".

    We reuse the value column as the name of the measure, so the measures will
    come from RADON_COMMON_MEASURES below.
    """

    # read each row, creating an object from it and adding it to the list
    obj_list = []
    for row in rows:
        # create a new object
        obj_list.append(model(**{
            "FIPS" : row["fips"],
            "County" : row["county"],
            "State" : "Colorado",
            "measure" : measure,
            "value": row["value"]
        }))

    # bulk insert all objects
    session.add_all(obj_list)

# columns in both county and tract sheets that we'll import as measures
RADON_COMMON_MEASURES = [
    "NTests", # Total Tests
    "NTestsover4", # Total Test over 4 pCi/L
    "PctOver4", # Percentage of Tests over 4 pCi/L
]

# metadata about each sheet in the radon spreadsheet, since there's both a
# county and tract sheet. we assume they're in the same order as this list.
RADON_SHEETS_META = [
    {
        "name": "CountyResults2005_2022",
        "model": RadonCounty,
        "fips_col": "CountyFIPSCode",
        "county_col": "CountyName",
        "measures": RADON_COMMON_MEASURES
    },
    {
        "name": "CensusTractResults2005_2022",
        "model": RadonTract,
        "fips_col": "CensusTract",
        "county_col": "CountyName",
        "measures": RADON_COMMON_MEASURES
    }
]

async def import_radon(excel_path, delete_before_import=True):
    """
    Imports the county and tract sheets within the radon data spreadsheet.

    Iterates over each sheet, zipping it with RADON_SHEETS_META to get
    metadata about the sheet. Then, for each column in RADON_SHEETS_META,
    it imports that column's value, using the column name as the measure name.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    # step 1. convert input excel using the column mapper

    # for each file, import it into the database
    async with async_session() as session:
        tqdm.write(f"* About to process {excel_path}...")

        # delete all existing entries in the target models
        if delete_before_import:
            for model in (RadonCounty, RadonTract):
                result = await session.execute(delete(model))
                tqdm.write(f" - Deleted {result.rowcount} from {model.__name__}")

        # read in the excel file via openpyxl
        try:
            wb = openpyxl.load_workbook(excel_path)

            # process each sheet, assuming they're in (county, tract) order
            for ws, sheet_meta in zip(wb.worksheets, RADON_SHEETS_META):
                # verify that we're looking at the sheet we expect
                if ws.title != sheet_meta['name']:
                    raise UnexpectedSheetException(
                        f"Expected sheet {sheet_meta['name']}, got {ws.title}"
                    )

                # create a dictionary of column names
                col_names = [
                    x[0].strip()
                    for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
                    if x[0] is not None
                ]
                col_indices = { col: idx for idx, col in enumerate(col_names) }

                # import each measure for the current sheet
                for col in sheet_meta['measures']:
                    tqdm.write(f"* Processing {ws.title}, {col}...")

                    # notes on input:
                    # - the FIPS column in the input doesn't include the leading 0
                    # - the 'PctOver4' percent column is 0-100, whereas percent
                    #   measures are 0-1.0 elsewhere, so we normalize it here
                    # - some tracts have no info, so we just skip it if there
                    #   isn't a value in the column

                    # read the column, normalizing sheet cols to internal names
                    rows = [
                        {
                            "fips": f"0{row[col_indices[sheet_meta['fips_col']]]}",
                            "county": row[col_indices[sheet_meta["county_col"]]],
                            "value": (
                                row[col_indices[col]]
                                if col != "PctOver4"
                                else row[col_indices[col]] / 100
                            )
                        }
                        for row in ws.iter_rows(min_row=2, values_only=True)
                        if row[col_indices[col]] is not None
                    ]
                    
                    await import_measure(
                        model=sheet_meta["model"],
                        measure=col,
                        rows=rows,
                        session=session
                    )

                    # commit session at the end
                    await session.commit()

        finally:
            wb.close()

@click.command()
@click.argument('excel-path', type=str)
def main(excel_path):
    asyncio.run(import_radon(excel_path))

if __name__ == '__main__':
    main()
