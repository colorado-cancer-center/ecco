#!/usr/bin/env python

import sys
import asyncio
import click
import openpyxl
import csv

from sqlalchemy.orm import sessionmaker
from sqlmodel import delete
from tqdm import tqdm

sys.path.append("/app")

from db import engine
from models.vaping import StateYouthVapingStats, VapingHealthRegion
from sqlalchemy.ext.asyncio import AsyncSession

SHEET_MEASURES_TO_QUESTION_KEYS = {
    "Percentage of students who have ever used an electronic vapor product": "youth_vaping_ever",
    "Percentage of students who used an electronic vapor product in the past 30 days": "youth_vaping_now",
    "ever": "adult_vaping_ever",
    "now": "adult_vaping_now",
}


def read_sheet(sheet_fp, sheet_name=None):
    """
    Given an Excel sheet, read the data from it and yield
    a dictionary for each row, where the keys are the column
    headers and the values are the cell values.
    """

    wb = openpyxl.load_workbook(sheet_fp)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    # get the header row
    header = [cell.value for cell in ws[1]]

    # get the data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(header, row))

def read_csv(csv_fp):
    """
    Given a CSV file pointer, read the data from it and yield
    a dictionary for each row, where the keys are the column
    headers and the values are the cell values.
    """

    for row in csv.DictReader(csv_fp):
        result = {k.strip(): v for k, v in row.items()}
        yield result

async def import_into_model(async_session, model, label, records):
    """
    Helper to insert records into model, clearing the model beforehand if requested.
    """
    async with async_session() as session:
        tqdm.write(f"* Processing {label}...")
        new_recs = [
            model(**rec)
            for rec in records
        ]
        session.add_all(new_recs)

        tqdm.write(f"* Inserted {len(new_recs)} records")
        tqdm.write("")

        await session.commit()


async def import_vaping_data(youth_regional_csv, adult_regional_sheet, delete_before_import=True):
    """
    Imports health-region vaping data from an Excel sheet into the database.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    if delete_before_import:
        async with async_session() as session:
            print("* Deleting existing vaping data...")
            for model in (StateYouthVapingStats, VapingHealthRegion):
                result = await session.execute(delete(model))
                print(f" - Deleted {result.rowcount} from {model.__name__}")
            print("")
            await session.commit()

    # ======================================================================
    # step 1. bring in youth data, which is combined HSR and state-level CSV
    # ======================================================================

    # the sheet has the following columns w/sample values (usage is below each):

    # Health Statistics Region: "Colorado"
    # - typically "Region 1", but state-level data is labeled "Colorado"
    # Year: "2023"
    # Health Topic: "Tobacco"
    # - unused
    # Health Measure: "Percentage of students who have ever used an electronic vapor product"
    # - maps to the measure key via SHEET_MEASURES_TO_QUESTION_KEYS
    # Demographic: "Total"
    # - unused
    # Population: "Total"
    # - unused
    # Percent (%): "30.00%"
    # - used as the value (divided by 100 to get a proportion)
    # Lower 95% Confidence Interval: "28.60%"
    # - unused
    # Upper 95% Confidence Interval: "31.40%"
    # - unused

    state_recs = []
    youth_hsr_recs = []

    for rec in read_csv(youth_regional_csv):
        if rec["Health Statistics Region"].strip() == "Colorado":
            state_recs.append(rec)
        else:
            youth_hsr_recs.append(rec)

    # first import the state-level youth data
    await import_into_model(
        async_session, StateYouthVapingStats, "youth Colorado-level vaping data",
        (
            {
                "measure_category": "Youth Vaping",
                "measure": rec["Health Measure"],
                "state_avg": float(rec["Percent (%)"].replace("%", "")) / 100.0,
                "us_avg": None,
                "source": f"Healthy Kids Colorado Survey, {rec['Year']}",
            }
            for rec in state_recs
        )
    )

    # then import the HSR-level youth data
    await import_into_model(
        async_session, VapingHealthRegion, "youth HSR-level vaping data",
        (
            {
                "hs_region": rec["Health Statistics Region"].split(" ")[-1],
                "State": "Colorado",
                "measure": SHEET_MEASURES_TO_QUESTION_KEYS[rec["Health Measure"]],
                "value": (
                    float(rec["Percent (%)"].replace("%", "")) / 100.0
                    if rec.get("Percent (%)", "").strip() != ""
                    else None
                ),
                "source": f"Healthy Kids Colorado Survey, {rec['Year']}"
            }
            for rec in youth_hsr_recs
        )
    )

    # ======================================================================
    # step 2. bring in adult HSR-level data, which is in a separate Excel sheet
    # ======================================================================

    # the sheet has the following columns w/sample values (usage is below each
    # column defn):

    # HSR: "HSR 1: Logan, Morgan, Phillips, Sedgwick, Washington, and Yuma Counties"
    # - parsed to get the HSR number
    # PCT: "23.2079"
    # - used as the value
    # 95 % LCI: "9.3387"
    # - unused
    # 95 % UCI: "37.0771"
    # - unused
    # Statistic: "ever"
    # - maps to the measure key via SHEET_MEASURES_TO_QUESTION_KEYS

    await import_into_model(
        async_session, VapingHealthRegion, "adult HSR-level vaping data",
        (
            {
                "hs_region": rec["HSR"].split(":")[0].replace("HSR ", ""),
                "State": "Colorado",
                "measure": SHEET_MEASURES_TO_QUESTION_KEYS[rec["Statistic"]],
                "value": rec["PCT"] / 100.0 if rec["PCT"] is not None else None,
            }
            for rec in read_sheet(adult_regional_sheet)
        )
    )

@click.command()
@click.argument('youth-regional-csv', type=click.File('r', encoding='utf-8-sig'))
@click.argument('adult-regional-sheet', type=click.File('rb'))
def main(youth_regional_csv, adult_regional_sheet):
    asyncio.run(import_vaping_data(
        youth_regional_csv, adult_regional_sheet
    ))

if __name__ == '__main__':
    main()
