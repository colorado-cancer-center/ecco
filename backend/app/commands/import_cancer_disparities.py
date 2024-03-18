#!/usr/bin/env python

import sys

sys.path.append("/app")

import click
import asyncio

from typing import get_type_hints

import openpyxl
from openpyxl.utils import get_column_letter

from tqdm import tqdm
from sqlmodel import delete, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import AsyncSession

from db import engine

from models import County
from models.disparity_index import (
    CancerDisparitiesIndex
)


class ModelNotProvidedException(Exception):
    """
    Raised when an input file is found that has no associated
    model class.
    """
    pass

async def import_measure(rows, measure_name, session):
    """
    Imports a table of measurements into 'model' in the database.

    If delete_before_import is True, all existing entries in the target model
    are deleted before importing. 
    """

    # read each row, creating an object from it and adding it to the list
    obj_list = []
    for row in rows:
        # look up the county object by its name, so we can get its FIPS
        county_result = await session.execute(
            select(County).where(County.label == row["county"])
        )
        chosen_county = county_result.scalars().first()

        # create a new object
        obj_list.append(CancerDisparitiesIndex(**{
            "FIPS" : chosen_county.us_fips,
            "County" : row["county"],
            "State" : "Colorado",
            "measure" : measure_name,
            "value": row["value"]
        }))

    # bulk insert all objects
    session.add_all(obj_list)

# maps columns in the input excel sheet to models
COLUMNS_TO_MEASURES = {
    "HNC Rank": "Head and Neck Cancer Index",
    "BC Rank": "Breast Cancer Index",
    "CRC Rank": "Colorectal Cancer Index",
    "LC Rank": "Lung Cancer Index"
}

async def import_disparities_index(excel_path, delete_before_import=True):
    """
    Imports the cancer disparity index file in these steps:
    1. converts wide table to a series of long tables in a similar format to CIF
    2. codes counties as FIPS identifiers
    3. inserts long tables into database.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    # step 1. convert input excel using the column mapper

    # for each file, import it into the database
    async with async_session() as session:
        tqdm.write(f"* About to process {excel_path}...")

        # delete all existing entries in the target model
        if delete_before_import:
            result = await session.execute(delete(CancerDisparitiesIndex))
            tqdm.write(f" - Deleted {result.rowcount} from {CancerDisparitiesIndex.__name__}")

        # read in the excel file via openpyxl
        try:
            wb = openpyxl.load_workbook(excel_path)
            # get the first sheet
            ws = wb.worksheets[0]

            # create a dictionary of column names
            col_names = [
                x[0].strip()
                for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
                if x[0] is not None
            ]
            col_indices = { col: idx for idx, col in enumerate(col_names) }

            # get the list of counties once, as we'll zip it with each value column
            counties = [
                x.value.strip()
                for x in ws[get_column_letter(col_indices["County"] + 1)][1:]
            ]

            for col, measure_name in COLUMNS_TO_MEASURES.items():
                tqdm.write(f"* Processing {col}...")

                # read the column
                rows = [
                    {
                        "county": county,
                        "measure": measure_name,
                        "value": row[col_indices[col]]
                    }
                    for county, row in zip(counties, ws.iter_rows(min_row=2, values_only=True))
                ]
                
                await import_measure(
                    rows, measure_name, session
                )
                    
                # tqdm.write(f"...insert done, committing...")
                # await session.commit()
                # tqdm.write(f"done!\n")

            # commit session at the end
            await session.commit()

        finally:
            wb.close()

@click.command()
@click.argument('excel-path', type=str)
def main(excel_path):
    asyncio.run(import_disparities_index(excel_path))

if __name__ == '__main__':
    main()
