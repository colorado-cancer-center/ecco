"""
Helpers and other bits for importing measure data from various sources into
ECCO's specific data model.

These are primarily used by the import_* commands in the commands/ directory.
"""

from typing import Callable, TypedDict, Optional
import openpyxl

from tqdm import tqdm
from models.base import BaseStatsModel
from sqlmodel import delete
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.asyncio import AsyncSession

from db import engine


# ==========================================================================
# === types that describe how to interpret the input sheets
# ==========================================================================

class MeasureMappingDict(TypedDict):
    """
    Defines the mapping between a column in the input sheet and the measure
    name that will be used in the database.

    The measure name is stored in the 'measure' column of the target model.
    """
    column: str
    measure: str

class MeasureCategoryDict(TypedDict):
    """
    Defines the mapping between columns in the input sheets and the database
    model (i.e., the measure category) that they'll populate. The model has
    a column 'measure' that specifies the different measures within that model.

    If `measures` is a list of strings, then each string is used as both the
    column name in the sheet and the measure name in the database. If it's a
    list of MeasureMappingDict objects, then each object specifies the Excel column
    name and the target measure name separately.
    """
    name: str
    model: BaseStatsModel
    fips_col: str
    county_col: str
    measures: list[MeasureMappingDict | str]

class MeasureRow(TypedDict):
    """
    Defines a single row of data to be imported. This is a mapping from the
    original sheet row, but normalized to the internal column names we use
    in import_measure().
    """
    fips: str
    county: str
    value: float


# ==========================================================================
# === exceptions that may occur during import
# ==========================================================================

class ModelNotProvidedException(Exception):
    """
    Raised when an input file is found that has no associated
    model class.
    """
    pass

class UnexpectedSheetException(Exception):
    """
    Raised when the sheet in the input file is not what was expected
    (i.e., it didn't match the name in UV_SHEETS_META for that index)
    """
    pass

# ==========================================================================
# === import helper
# ==========================================================================

async def _import_measure(model, measure, rows: list[MeasureRow], session):
    """
    Imports a table of measurements into 'model' in the database.

    Note that the "rows" object is mapped from the original sheet row in
    import_measures_from_sheet() below. We expect that each row object has the
    following keys: "fips", "county", and "value".
    """

    # read each row, creating an object from it and adding it to the list
    obj_list = []
    for row in rows:
        # create a new object
        obj_list.append(model(**{
            "FIPS" : row["fips"],
            "County" : row["county"],
            "State" : "Colorado",
            "measure" : measure,
            "value": row["value"]
        }))

    # bulk insert all objects
    session.add_all(obj_list)

def _zero_pad(fips: str) -> str:
    """
    We presume that all the FIPS codes are for Colorado, i.e. should start with
    '08'. This method normalizes the FIPS code to ensure that it starts with '0'
    if needed.

    If the input FIPS code is not a CO FIPS code, this raises a ValueError.
    """
    result = (
        f"0{fips}"
        if not isinstance(fips, str) or fips[0] != '0' else
        fips
    )
    
    # check that it's now 08 and throw an error if not
    if result[0:2] != '08':
        raise ValueError(f"FIPS code {fips} doesn't start with '08'")
    return result

async def import_measures_from_sheets(
    excel_path,
    sheets_meta: list[MeasureCategoryDict],
    delete_before_import=True,
    value_mapper: Optional[Callable[[float], float]]=None,
    normalize_fips=True
):
    """
    Imports multiple sheets within the input data spreadsheet as measure categories.
    
    The input is presumed to be an Excel file where each sheet corresponds to a
    different measure category. This is most often a different geographic level
    (e.g., county, tract), but could be any other categorization as long as each
    sheet is a measure category and its measures are located within the sheet as
    individual columns.

    The sheets_meta argument is a list of MeasureCategoryDict objects that describe
    how to interpret each sheet. We enforce that the sheets are in the same order
    as the list, and that the sheet names match the 'name' field in each
    MeasureCategoryDict.
    """

    async_session = sessionmaker(
        engine, expire_on_commit=False, class_=AsyncSession
    )

    # step 1. convert input excel using the column mapper

    # for each file, import it into the database
    async with async_session() as session:
        tqdm.write(f"* About to process {excel_path}...")

        # delete all existing entries in the target models
        if delete_before_import:
            for model in (sheet_meta['model'] for sheet_meta in sheets_meta):
                result = await session.execute(delete(model))
                tqdm.write(f" - Deleted {result.rowcount} from {model.__name__}")

        # read in the excel file via openpyxl
        try:
            wb = openpyxl.load_workbook(excel_path)

            # process each sheet, assuming they're in (county, tract) order
            for ws, sheet_meta in zip(wb.worksheets, sheets_meta):
                # verify that we're looking at the sheet we expect
                if ws.title != sheet_meta['name']:
                    raise UnexpectedSheetException(
                        f"Expected sheet {sheet_meta['name']}, got {ws.title}"
                    )

                # create a dictionary of column names
                col_names = [
                    x[0].strip()
                    for x in ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)
                    if x[0] is not None
                ]
                col_indices = { col: idx for idx, col in enumerate(col_names) }

                # import each measure for the current sheet
                for mapping in sheet_meta['measures']:
                    if isinstance(mapping, str):
                        # if it's just a string, use it for both column and measure
                        col = mapping
                        measure = mapping
                    else:
                        col = mapping['column']
                        measure = mapping['measure']
                        
                    tqdm.write(f"* Processing {ws.title}, {col}...")

                    # read the column, normalizing sheet cols to internal names
                    try:
                        rows = [
                            {
                                "fips": (
                                    _zero_pad(row[col_indices[sheet_meta['fips_col']]])
                                    if normalize_fips else
                                    row[col_indices[sheet_meta['fips_col']]]
                                ),
                                "county": row[col_indices[sheet_meta["county_col"]]],
                                "value": (
                                    value_mapper(row[col_indices[col]])
                                    if value_mapper is not None else
                                    row[col_indices[col]]
                                )
                            }
                            for row in ws.iter_rows(min_row=2, values_only=True)
                            if row[col_indices[col]] is not None
                        ]
                    except KeyError as e:
                        raise KeyError(
                            f"Column {e} not found in sheet '{ws.title}'; "
                            f"available columns: {col_names}"
                        ) from e
                    
                    await _import_measure(
                        model=sheet_meta["model"],
                        measure=measure,
                        rows=rows,
                        session=session
                    )

                    # commit session at the end
                    await session.commit()

        finally:
            wb.close()
